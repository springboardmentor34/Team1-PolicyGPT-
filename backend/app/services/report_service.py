import io
from typing import List
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from app.models.models import Policy, Scheme

def generate_policies_pdf(policies: List[Policy]) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    story = []

    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontSize=20,
        leading=24,
        textColor=colors.HexColor('#003366'),
        spaceAfter=12
    )
    story.append(Paragraph("PolicyGPT - Official Government Policy Intelligence Report", title_style))
    story.append(Spacer(1, 12))

    data = [["Code", "Policy Title", "Category", "Ministry", "Status", "State"]]
    for p in policies:
        data.append([
            p.code,
            p.title[:35] + ("..." if len(p.title) > 35 else ""),
            p.category,
            p.ministry[:25] + ("..." if len(p.ministry) > 25 else ""),
            p.status,
            p.state
        ])

    table = Table(data, colWidths=[80, 150, 90, 110, 60, 60])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8F9FA')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))
    story.append(table)

    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()

def generate_schemes_excel(schemes: List[Scheme]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Public Schemes Report"

    headers = ["ID", "Code", "Scheme Name", "Category", "Financial Assistance", "Budget Allocated (₹)", "Status", "Target Group", "Deadline"]
    ws.append(headers)

    # Style header row
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="003366", end_color="003366", fill_type="solid")

    for s in schemes:
        ws.append([
            s.id,
            s.code,
            s.name,
            s.category,
            s.financial_assistance or "N/A",
            float(s.budget_allocated) if s.budget_allocated else 0.0,
            s.status,
            s.target_group or "N/A",
            str(s.deadline) if s.deadline else "N/A"
        ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def generate_department_pdf(dept_name: str, policies: List[Policy], schemes: List[Scheme]) -> bytes:

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    story = []

    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontSize=18,
        leading=22,
        textColor=colors.HexColor('#003366'),
        spaceAfter=6
    )
    story.append(Paragraph(f"PolicyGPT — Department Intelligence Report: {dept_name}", title_style))
    story.append(Spacer(1, 10))

    story.append(Paragraph("<b>Department Policies Overview</b>", styles['Heading2']))
    story.append(Spacer(1, 6))

    p_data = [["Code", "Policy Title", "Category", "Status", "Views"]]
    for p in policies:
        p_data.append([
            p.code,
            p.title[:40] + ("..." if len(p.title) > 40 else ""),
            p.category,
            p.status,
            str(p.view_count or 0)
        ])

    if len(p_data) > 1:
        p_table = Table(p_data, colWidths=[90, 220, 100, 80, 50])
        p_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ]))
        story.append(p_table)
    else:
        story.append(Paragraph("No policies registered under this department.", styles['Normal']))

    story.append(Spacer(1, 14))
    story.append(Paragraph("<b>Department Public Schemes Overview</b>", styles['Heading2']))
    story.append(Spacer(1, 6))

    s_data = [["Code", "Scheme Name", "Category", "Financial Assistance", "Status"]]
    for s in schemes:
        s_data.append([
            s.code,
            s.name[:35] + ("..." if len(s.name) > 35 else ""),
            s.category,
            (s.financial_assistance or "N/A")[:25],
            s.status
        ])

    if len(s_data) > 1:
        s_table = Table(s_data, colWidths=[90, 200, 100, 90, 60])
        s_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D97706')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ]))
        story.append(s_table)
    else:
        story.append(Paragraph("No public schemes registered under this department.", styles['Normal']))

    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()

def generate_analytics_pdf(summary_data: dict) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    story = []

    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontSize=20,
        leading=24,
        textColor=colors.HexColor('#003366'),
        spaceAfter=10
    )
    story.append(Paragraph("PolicyGPT — Executive System Analytics Report", title_style))
    story.append(Spacer(1, 10))

    users = summary_data.get("users", {})
    policies = summary_data.get("policies", {})
    schemes = summary_data.get("schemes", {})

    metrics = [
        ["Metric", "Value"],
        ["Total Registered Users", str(users.get("total", 0))],
        ["Citizens", str(users.get("citizens", 0))],
        ["Government Officials", str(users.get("officials", 0))],
        ["Researchers", str(users.get("researchers", 0))],
        ["Organizations", str(users.get("organizations", 0))],
        ["Total Policies", str(policies.get("total", 0))],
        ["Published Policies", str(policies.get("published", 0))],
        ["Pending Policies", str(policies.get("pending", 0))],
        ["Total Schemes", str(schemes.get("total", 0))],
        ["Active Schemes", str(schemes.get("active", 0))]
    ]

    table = Table(metrics, colWidths=[250, 250])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(table)

    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()

